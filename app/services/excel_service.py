"""
Excel出力サービス

世話記録、診療記録などのExcel出力機能を提供します。

Requirements: Requirement 7.5, Requirement 9.4
Context7: /openpyxl/openpyxl
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy.orm import Session

from app.models.care_log import CareLog
from app.services.medical_report_service import get_medical_summary_rows
from app.utils.i18n import tj


def generate_care_log_excel(
    db: Session,
    start_date: date,
    end_date: date,
    animal_id: int | None = None,
    locale: str = "ja",
) -> bytes:
    """
    世話記録のExcelファイルを生成

    Args:
        db: データベースセッション
        start_date: 開始日
        end_date: 終了日
        animal_id: 猫のID（指定時は特定の猫のみ）

    Returns:
        bytes: Excel形式のバイト列

    Example:
        >>> from datetime import date
        >>> excel_data = generate_care_log_excel(
        ...     db, date(2024, 11, 1), date(2024, 11, 30)
        ... )
        >>> with open("care_logs.xlsx", "wb") as f:
        ...     f.write(excel_data)
    """
    # 世話記録を取得（実際の記録日でフィルタリング）
    query = db.query(CareLog).filter(
        CareLog.log_date >= start_date,
        CareLog.log_date <= end_date,
    )

    # 個別猫の場合はフィルター
    if animal_id:
        query = query.filter(CareLog.animal_id == animal_id)

    records = query.order_by(CareLog.log_date.desc(), CareLog.created_at.desc()).all()

    # Excelワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = tj("sheet_names.care_logs", locale=locale)

    # ヘッダースタイル
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # ヘッダー行（多言語化）
    headers = [
        tj("headers.created_at", locale=locale),
        tj("headers.log_date", locale=locale),
        tj("headers.animal_id", locale=locale),
        tj("headers.animal_name", locale=locale),
        tj("headers.time_slot", locale=locale),
        tj("headers.appetite", locale=locale),
        tj("headers.energy", locale=locale),
        tj("headers.urination", locale=locale),
        tj("headers.cleaning", locale=locale),
        tj("headers.recorder_id", locale=locale),
        tj("headers.recorder_name", locale=locale),
        tj("headers.memo", locale=locale),
        tj("headers.ip_address", locale=locale),
        tj("headers.device_tag", locale=locale),
        tj("headers.from_paper", locale=locale),
        tj("headers.last_updated_at", locale=locale),
        tj("headers.last_updated_by", locale=locale),
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # データ行
    for row_num, record in enumerate(records, 2):
        # 猫名を取得
        animal_name = ""
        if record.animal:
            animal_name = record.animal.name or tj(
                "animal.no_name", locale=locale, id=record.animal_id
            )
        else:
            animal_name = tj("animal.no_name", locale=locale, id=record.animal_id)

        row_data = [
            record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            record.log_date.strftime("%Y-%m-%d"),
            record.animal_id,
            animal_name,
            tj(f"time_slots.{record.time_slot}", locale=locale),
            record.appetite,
            record.energy,
            tj(
                "boolean.yes" if record.urination else "boolean.no",
                locale=locale,
            ),
            tj(
                "boolean.yes" if record.cleaning else "boolean.no",
                locale=locale,
            ),
            record.recorder_id or "",
            record.recorder_name,
            record.memo or "",
            record.ip_address or "",
            record.device_tag or "",
            tj(
                "boolean.yes" if record.from_paper else "boolean.no",
                locale=locale,
            ),
            record.last_updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            record.last_updated_by or "",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            # 中央揃え（一部のカラム）
            if col_num in [5, 6, 7, 8, 9, 15]:  # 時点、食欲、元気、排尿、清掃、紙記録
                cell.alignment = Alignment(horizontal="center")

    # 列幅を自動調整
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        # 最大幅を設定（見やすさのため）
        max_width = 15
        if col_num == 12:  # メモ列
            max_width = 30
        elif col_num in [1, 2, 16]:  # 日時列
            max_width = 20

        ws.column_dimensions[column_letter].width = max_width

    # フリーズペイン（ヘッダー行を固定）
    ws.freeze_panes = "A2"

    # Excelファイルをバイト列として出力
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def generate_report_excel(
    db: Session,
    report_type: str,
    start_date: date,
    end_date: date,
    animal_id: int | None = None,
    locale: str = "ja",
) -> bytes:
    """
    帳票Excelファイルを生成（日報・週報・月次集計・個別帳票）

    Args:
        db: データベースセッション
        report_type: 帳票種別（daily/weekly/monthly/individual）
        start_date: 開始日
        end_date: 終了日
        animal_id: 猫のID（個別帳票の場合のみ必須）
        locale: ロケール（ja/en）

    Returns:
        bytes: Excel形式のバイト列

    Raises:
        ValueError: 不正な帳票種別の場合、または個別帳票で猫IDが未指定の場合

    Example:
        >>> from datetime import date
        >>> excel_data = generate_report_excel(
        ...     db, "daily", date(2024, 11, 1), date(2024, 11, 30)
        ... )
        >>> with open("report.xlsx", "wb") as f:
        ...     f.write(excel_data)
    """
    # 帳票種別のバリデーション
    valid_types = ["daily", "weekly", "monthly", "individual", "medical_summary"]
    if report_type not in valid_types:
        raise ValueError(
            f"不正な帳票種別です: {report_type}。有効な値: {', '.join(valid_types)}"
        )

    # 個別帳票の場合は猫IDが必須
    if report_type == "individual" and not animal_id:
        raise ValueError("個別帳票の生成には猫IDが必要です")

    if report_type == "medical_summary":
        return generate_medical_summary_excel(
            db, start_date, end_date, animal_id, locale
        )

    # 世話記録Excelを生成（全帳票種別で共通）
    return generate_care_log_excel(db, start_date, end_date, animal_id, locale)


def generate_medical_summary_excel(
    db: Session,
    start_date: date,
    end_date: date,
    animal_id: int | None = None,
    locale: str = "ja",
) -> bytes:
    """診療記録（利益計算用）Excelファイルを生成"""

    rows, _totals = get_medical_summary_rows(
        db=db, start_date=start_date, end_date=end_date, animal_id=animal_id
    )

    wb = Workbook()
    ws = wb.active
    ws.title = tj("sheet_names.medical_records", locale=locale)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        tj("headers.medical_record_id", locale=locale),
        tj("headers.medical_date", locale=locale),
        tj("headers.animal_id", locale=locale),
        tj("headers.animal_name", locale=locale),
        tj("headers.medical_action_name", locale=locale),
        tj("headers.dosage", locale=locale),
        tj("headers.dosage_unit", locale=locale),
        tj("headers.cost_price", locale=locale),
        tj("headers.selling_price", locale=locale),
        tj("headers.procedure_fee", locale=locale),
        tj("headers.billing_amount", locale=locale),
        tj("headers.currency", locale=locale),
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, row in enumerate(rows, 2):
        ws.cell(row=row_num, column=1).value = row.medical_record_id
        ws.cell(row=row_num, column=2).value = row.medical_date.strftime("%Y-%m-%d")
        ws.cell(row=row_num, column=3).value = row.animal_id
        ws.cell(row=row_num, column=4).value = row.animal_name
        ws.cell(row=row_num, column=5).value = row.medical_action_name or ""
        ws.cell(row=row_num, column=6).value = row.dosage
        ws.cell(row=row_num, column=7).value = row.dosage_unit or ""
        ws.cell(row=row_num, column=8).value = (
            float(row.cost_price) if row.cost_price is not None else None
        )
        ws.cell(row=row_num, column=9).value = (
            float(row.selling_price) if row.selling_price is not None else None
        )
        ws.cell(row=row_num, column=10).value = (
            float(row.procedure_fee) if row.procedure_fee is not None else None
        )
        ws.cell(row=row_num, column=11).value = (
            float(row.billing_amount) if row.billing_amount is not None else None
        )
        ws.cell(row=row_num, column=12).value = row.currency or ""

        # Align numeric-ish columns
        for col_num in [1, 2, 3, 6, 8, 9, 10, 11, 12]:
            ws.cell(row=row_num, column=col_num).alignment = Alignment(
                horizontal="center"
            )

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_width = 18
        if col_num in [4, 5]:
            max_width = 24
        ws.column_dimensions[column_letter].width = max_width

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
