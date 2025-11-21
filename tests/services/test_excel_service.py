"""
Excelサービスのテスト（多言語対応）

Requirements: Requirement 9, Requirement 25
"""

from __future__ import annotations

from datetime import date

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.animal import Animal
from app.models.care_log import CareLog
from app.services.excel_service import (
    generate_care_log_excel,
    generate_report_excel,
)


class TestGenerateCareLogExcel:
    """ケアログExcel生成のテスト"""

    def test_generate_care_log_excel_japanese(
        self, test_db: Session, test_care_logs: list[CareLog]
    ):
        """正常系: 日本語でケアログExcelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_care_log_excel(
            test_db, start_date, end_date, locale="ja"
        )

        # Then
        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Excelファイルとして読み込める
        from io import BytesIO

        wb = load_workbook(BytesIO(excel_bytes))
        assert wb is not None
        assert "世話記録" in wb.sheetnames

    def test_generate_care_log_excel_english(
        self, test_db: Session, test_care_logs: list[CareLog]
    ):
        """正常系: 英語でケアログExcelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_care_log_excel(
            test_db, start_date, end_date, locale="en"
        )

        # Then
        assert excel_bytes is not None
        from io import BytesIO

        wb = load_workbook(BytesIO(excel_bytes))
        assert "Care Logs" in wb.sheetnames

    def test_generate_care_log_excel_no_data(self, test_db: Session):
        """境界値: データなしでも正常にExcel生成できる"""
        # Given
        start_date = date(2025, 1, 1)
        end_date = date(2025, 1, 31)

        # When
        excel_bytes = generate_care_log_excel(
            test_db, start_date, end_date, locale="ja"
        )

        # Then
        assert excel_bytes is not None


class TestGenerateReportExcel:
    """帳票Excel生成のテスト"""

    def test_generate_daily_report_excel_japanese(
        self, test_db: Session, test_care_logs: list[CareLog]
    ):
        """正常系: 日本語で日報Excelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_report_excel(
            test_db, "daily", start_date, end_date, locale="ja"
        )

        # Then
        assert excel_bytes is not None

    def test_generate_weekly_report_excel_english(
        self, test_db: Session, test_care_logs: list[CareLog]
    ):
        """正常系: 英語で週報Excelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_report_excel(
            test_db, "weekly", start_date, end_date, locale="en"
        )

        # Then
        assert excel_bytes is not None

    def test_generate_monthly_report_excel(
        self, test_db: Session, test_care_logs: list[CareLog]
    ):
        """正常系: 月次集計Excelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_report_excel(
            test_db, "monthly", start_date, end_date, locale="ja"
        )

        # Then
        assert excel_bytes is not None

    def test_generate_individual_report_excel_with_animal(
        self,
        test_db: Session,
        test_animal: Animal,
        test_care_logs: list[CareLog],
    ):
        """正常系: 個別帳票Excelを生成できる"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When
        excel_bytes = generate_report_excel(
            test_db,
            "individual",
            start_date,
            end_date,
            animal_id=test_animal.id,
            locale="ja",
        )

        # Then
        assert excel_bytes is not None

    def test_generate_report_excel_invalid_type(self, test_db: Session):
        """異常系: 不正な帳票種別で例外発生"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When/Then
        import pytest

        with pytest.raises(ValueError):
            generate_report_excel(test_db, "invalid", start_date, end_date, locale="ja")

    def test_generate_individual_report_excel_missing_animal_id(self, test_db: Session):
        """異常系: 個別帳票で猫ID未指定で例外発生"""
        # Given
        start_date = date(2024, 11, 1)
        end_date = date(2024, 11, 30)

        # When/Then
        import pytest

        with pytest.raises(ValueError):
            generate_report_excel(
                test_db, "individual", start_date, end_date, locale="ja"
            )
